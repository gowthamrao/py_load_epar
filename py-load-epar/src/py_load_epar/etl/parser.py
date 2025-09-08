import logging
from pathlib import Path
from typing import Any, Dict, Iterator

import openpyxl

logger = logging.getLogger(__name__)


def _snake_case(s: str) -> str:
    """
    Converts a string to a valid snake_case identifier.
    Example: 'Marketing Authorisation Holder' -> 'marketing_authorisation_holder'.
    """
    import re
    if not isinstance(s, str):
        return ""
    # Replace known separators with underscore
    s = re.sub(r"[ -/]", "_", s)
    # Handle camelCase by inserting underscore before uppercase letters
    s = re.sub(r"(?<=\w)([A-Z])", r"_\1", s)
    # Remove any characters that are not alphanumeric or underscore
    s = re.sub(r"[^a-zA-Z0-9_]", "", s)
    return s.lower()


from typing import IO, Union


def parse_ema_excel_file(
    file_source: Union[Path, IO[bytes]]
) -> Iterator[Dict[str, Any]]:
    """
    Parses an EMA Excel file and yields each row as a dictionary.

    This function is memory-efficient as it reads the workbook in read-only
    mode and yields rows one by one using a generator. It also dynamically
    maps columns based on the header row, making it resilient to changes in
    column order. The header names are converted to snake_case to align with
    Pydantic model field names.

    Args:
        file_source: The source of the Excel data. Can be a Path object to a
                     local file or a file-like object (e.g., io.BytesIO) in
                     binary mode.

    Yields:
        A dictionary representing a single row from the Excel file.
    """
    logger.info(f"Starting to parse Excel file from source: {type(file_source)}")
    try:
        # openpyxl.load_workbook can accept either a filename (Path) or a
        # file-like object directly.
        workbook = openpyxl.load_workbook(filename=file_source, read_only=True)
        sheet = workbook.active

        if sheet is None:
            raise ValueError("Excel file contains no active sheets.")

        # Read the header row and convert column names to snake_case
        headers = [_snake_case(cell.value) for cell in sheet[1]]
        logger.debug(f"Parsed Excel headers: {headers}")

        # Yield each subsequent row as a dictionary mapped by the headers
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            yield dict(zip(headers, row))

    except Exception as e:
        logger.error(f"Failed to parse Excel file from {type(file_source)}. Error: {e}")
        # Re-raise the exception to be handled by the calling orchestrator
        raise

    logger.info(f"Finished parsing Excel file from {type(file_source)}")
